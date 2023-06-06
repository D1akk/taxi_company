from typing import Any, Dict
from django.forms.models import BaseModelForm
from django.http import HttpResponse
from django.shortcuts import render
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView, FormView
from django.urls import reverse_lazy
from .models import *
from django.contrib.auth.views import LoginView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth import login, get_user_model
from .forms import *
from django.db.models.signals import post_save
from django.dispatch import receiver
from openpyxl import Workbook


def index(request):
    return render(request, 'record_transport/index.html')

class CustomLoginView(LoginView):
    template_name = 'record_transport/login.html'
    fields = '__all__'
    redirect_authenticated_user = True

    def get_success_url(self):
        return reverse_lazy('home')
  

User = get_user_model()


class RegisterAdminView(LoginRequiredMixin, CreateView):
    model = User
    form_class = RegisterAdminForm
    template_name = 'record_transport/register_admin.html'
    success_url = reverse_lazy('home')


class RegisterDriverView(LoginRequiredMixin, CreateView):
    model = User
    form_class = RegisterDriverForm
    template_name = 'record_transport/register_driver.html'
    success_url = reverse_lazy('driver-list')


class DriverListView(LoginRequiredMixin, ListView):
    model = Driver
    context_object_name = 'drivers'
    

class DriverDetailView(LoginRequiredMixin, DeleteView):
    model = Driver
    context_object_name = 'task'
    template_name = 'record_transport/driver.html'


class CarListView(LoginRequiredMixin, ListView):
    model = Car
    context_object_name = 'cars'


class TaskListView(LoginRequiredMixin, ListView):
    model = Task
    context_object_name = 'tasks'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tasks'] = context['tasks']#.filter(driver=self.request.user)
        context['count'] = context['tasks'].filter(complete=False).count()
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and (not user.is_staff):
            queryset = queryset.filter(driver=user)
        return queryset


class TaskDetailView(LoginRequiredMixin, DetailView):
    model = Task
    context_object_name = 'task'
    template_name = 'record_transport/task.html'


class TaskCreateView(LoginRequiredMixin, CreateView):
    model = Task
    form_class = TaskForm
    template_name = 'record_transport/task_create.html'
    success_url = reverse_lazy('tasks')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    

class TaskUpdateView(LoginRequiredMixin, UpdateView):
    model = Task
    form_class = TaskForm
    template_name = 'record_transport/task_update.html'
    success_url = reverse_lazy('tasks')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    

class TaskDeleteView(LoginRequiredMixin, DeleteView):
    model = Task
    context_object_name = 'task'
    success_url = reverse_lazy('tasks')


def generate_waybill(request, pk):
    task = Task.objects.get(id=pk)
    driver = task.driver
    car = Car.objects.get(driver_id=driver.pk)

    workbook = Workbook()
    sheet = workbook.active

    sheet.title = 'Путевой лист'
    sheet['B1'] = 'Название'
    sheet.cell(row=2, column=2).value = task.title
    
    sheet['B4'] = 'ФИО водителя'
    sheet.cell(row=5, column=2).value = driver.first_name+' '+driver.middle_name+' '+driver.last_name
    
    sheet['B7'] = 'Номер водительского удостоверения'
    sheet.cell(row=8, column=2).value = driver.license_number

    sheet['B10'] = 'Марка, модель автомобиля'
    sheet.cell(row=11, column=2).value = car.mark+' '+car.model
    
    sheet['B13'] = 'Государственный номерной знак'
    sheet.cell(row=14, column=2).value = car.number
    
    sheet['B16'] = 'Сведения о перевозке'
    sheet.cell(row=17, column=2).value = task.title+' '+str(task.date)+' '+str(task.time)

    sheet['B19'] = 'Показания одометра при выезде'
    sheet.cell(row=20, column=2).value = task.fuel_before

    sheet['B21'] = 'Показания одометра по приезду'
    sheet.cell(row=22, column=2).value = task.fuel_after
    
    sheet['B23'] = 'Пройденный километраж'
    sheet.cell(row=24, column=2).value = str(task.fuel_after - task.fuel_before)+" км."

    sheet['B25'] = 'Расход топлива'
    sheet.cell(row=26, column=2).value = str((task.fuel_after - task.fuel_before)*car.pokazatel)+" л."

    response = HttpResponse(content_type='application/vd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment;filename=putovoi_list_{task.id}_.xlsx'

    workbook.save(response)

    return response
        